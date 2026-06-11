from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from rbd.models import RbdServicio


JORNADAS = {
    "A": {
        "horario": "LMMJV 8:00 - 20:00 / Sabado 8:00 - 20:00",
        "descripcion": "Jornada Escolar Completa",
    },
    "B": {
        "horario": "LMMJV 8:00 - 16:00",
        "descripcion": "Sin Jornada Escolar Completa",
    },
    "C": {
        "horario": "LMMJV 8:00 - 18:00",
        "descripcion": "Zona rural de baja electricidad",
    },
    "D": {
        "horario": "LMMJV 8:00 - 23:30",
        "descripcion": "Jornada vespertina adultos",
    },
    "E": {
        "horario": "Horarios especiales de funcionamiento segun lo informado por MINEDUC y SUBTEL para ese EES",
        "descripcion": "Situaciones Especiales",
    },
}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_map(row):
    return {_clean(value).upper(): index for index, value in enumerate(row)}


def _value(row, headers, name):
    index = headers.get(name.strip().upper())
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


class Command(BaseCommand):
    help = "Importa jornada horaria por RBD desde Excel."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Ruta al Excel BBDD Proyecto CPE2030 Jornada Horaria.xlsx.")

    def handle(self, *args, **options):
        source = Path(options["path"])
        if not source.exists():
            raise CommandError(f"No existe el archivo: {source}")

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("Falta openpyxl. Reinstala dependencias del proyecto.") from exc

        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet = workbook["Jornada"] if "Jornada" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = _header_map(next(rows))

        updated = 0
        created = 0
        skipped = 0

        with transaction.atomic():
            for row in rows:
                raw_rbd = _value(row, headers, "RBD")
                raw_categoria = _value(row, headers, "JORNADA").upper()
                if not raw_rbd or not raw_categoria:
                    skipped += 1
                    continue
                try:
                    rbd = int(float(raw_rbd))
                except ValueError:
                    skipped += 1
                    continue

                categoria = raw_categoria[:1]
                detalle = JORNADAS.get(categoria)
                if not detalle:
                    skipped += 1
                    continue

                servicio, was_created = RbdServicio.objects.get_or_create(rbd=rbd)
                datos = dict(servicio.datos or {})
                datos["jornada_excel"] = {
                    "categoria": categoria,
                    "establecimiento": _value(row, headers, " ESTABLECIMIENTO EES") or _value(row, headers, "ESTABLECIMIENTO EES"),
                    "region": _value(row, headers, "REGION") or _value(row, headers, "REGION") or _value(row, headers, "REGIÃ“N"),
                    "comuna": _value(row, headers, "COMUNA"),
                    "direccion": _value(row, headers, "DIRECCION EES"),
                }
                servicio.jornada_categoria = categoria
                servicio.jornada_horario = detalle["horario"]
                servicio.jornada_descripcion = detalle["descripcion"]
                servicio.datos = datos
                servicio.save(
                    update_fields=[
                        "jornada_categoria",
                        "jornada_horario",
                        "jornada_descripcion",
                        "datos",
                        "actualizado_en",
                    ]
                )
                updated += 1
                created += int(was_created)

        self.stdout.write(
            self.style.SUCCESS(
                f"Jornadas actualizadas: {updated}. Servicios creados: {created}. Filas omitidas: {skipped}."
            )
        )