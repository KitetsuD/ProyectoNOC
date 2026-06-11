import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from rbd.models import RbdContacto, RbdServicio


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_map(row):
    return {_clean(value): index for index, value in enumerate(row)}


def _value(row, headers, name):
    index = headers.get(name)
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


def _rbd_from_text(value):
    match = re.search(r"RBD[_\s-]*(\d+)", _clean(value).upper())
    if match:
        return int(match.group(1))
    return None


class Command(BaseCommand):
    help = "Importa BPI y contactos RBD desde archivos Excel."

    def add_arguments(self, parser):
        parser.add_argument("--bpi", help="Ruta al Excel BPI-RBD.xlsx.")
        parser.add_argument("--contactos", help="Ruta al Excel de contactos.")

    def handle(self, *args, **options):
        if not options.get("bpi") and not options.get("contactos"):
            raise CommandError("Debes entregar --bpi, --contactos o ambos.")

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("Falta openpyxl. Reinstala dependencias del proyecto.") from exc

        total_bpi = 0
        total_contactos = 0
        servicios_creados = 0

        with transaction.atomic():
            if options.get("bpi"):
                updated, created = self._import_bpi(openpyxl, Path(options["bpi"]))
                total_bpi += updated
                servicios_creados += created
            if options.get("contactos"):
                updated, created = self._import_contactos(openpyxl, Path(options["contactos"]))
                total_contactos += updated
                servicios_creados += created

        self.stdout.write(
            self.style.SUCCESS(
                f"BPI actualizados: {total_bpi}. Contactos actualizados: {total_contactos}. "
                f"Servicios creados: {servicios_creados}."
            )
        )

    def _import_bpi(self, openpyxl, source):
        if not source.exists():
            raise CommandError(f"No existe el archivo BPI: {source}")
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet = workbook["BPI_RBD"] if "BPI_RBD" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = _header_map(next(rows))
        updated = 0
        created = 0

        for row in rows:
            codigo = _value(row, headers, "Codigo del Servicio") or _value(row, headers, "Código del Servicio")
            rbd = _rbd_from_text(codigo)
            if not rbd:
                continue

            servicio, was_created = RbdServicio.objects.get_or_create(rbd=rbd)
            datos = dict(servicio.datos or {})
            datos["bpi_excel"] = {
                "nombre": _value(row, headers, "Nombre"),
                "codigo_servicio": codigo,
                "numero_bpi": _value(row, headers, "Numero de BPI") or _value(row, headers, "Número de BPI"),
                "estado": _value(row, headers, "Estado"),
                "zona": _value(row, headers, "Zona"),
            }
            servicio.bpi = _value(row, headers, "Nombre") or datos["bpi_excel"]["numero_bpi"]
            if not servicio.zona:
                servicio.zona = _value(row, headers, "Zona")
            if not servicio.direccion:
                servicio.direccion = _value(row, headers, "Ubicacion de destino") or _value(row, headers, "Ubicación de destino")
            servicio.datos = datos
            servicio.save(update_fields=["bpi", "zona", "direccion", "datos", "actualizado_en"])
            updated += 1
            created += int(was_created)

        return updated, created

    def _import_contactos(self, openpyxl, source):
        if not source.exists():
            raise CommandError(f"No existe el archivo de contactos: {source}")
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet = workbook["9.414 EE"] if "9.414 EE" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = None
        updated = 0
        created = 0

        for row in rows:
            first_cell = _clean(row[0] if row else "")
            if first_cell.upper() == "RBD":
                headers = _header_map(row)
                break

        if not headers:
            raise CommandError("No se encontro la fila de encabezados en el Excel de contactos.")

        for row in rows:
            raw_rbd = _value(row, headers, "RBD")
            if not raw_rbd:
                continue
            try:
                rbd = int(float(raw_rbd))
            except ValueError:
                continue

            servicio, was_created = RbdServicio.objects.get_or_create(rbd=rbd)
            created += int(was_created)
            contacts = [
                {
                    "orden": 1,
                    "nombre": _value(row, headers, "Director"),
                    "telefono": _value(row, headers, "Fono Director") or _value(row, headers, "Fono EE"),
                    "celular": _value(row, headers, "Celular EE"),
                    "email": _value(row, headers, "Email  Direc") or _value(row, headers, "Email EE"),
                    "cargo": "Director",
                    "fuente": "Excel contactos",
                },
                {
                    "orden": 2,
                    "nombre": _value(row, headers, "Coordinador Informatica") or _value(row, headers, "Coordinador Informática"),
                    "telefono": _value(row, headers, "Celular CI"),
                    "celular": "",
                    "email": _value(row, headers, "Email CI"),
                    "cargo": "Coordinador Informatica",
                    "fuente": "Excel contactos",
                },
                {
                    "orden": 3,
                    "nombre": _value(row, headers, "JEFE UTP"),
                    "telefono": "",
                    "celular": _value(row, headers, "Celular UTP"),
                    "email": _value(row, headers, "Email UTP"),
                    "cargo": "Jefe UTP",
                    "fuente": "Excel contactos",
                },
            ]
            for contact in contacts:
                orden = contact.pop("orden")
                if not any(contact.get(field) for field in ("nombre", "telefono", "celular", "email")):
                    continue
                RbdContacto.objects.update_or_create(
                    servicio=servicio,
                    orden=orden,
                    defaults=contact,
                )
                updated += 1

        return updated, created
