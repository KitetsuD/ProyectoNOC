from datetime import date, datetime
from pathlib import Path
import re
import unicodedata

from django.db import transaction
from django.utils import timezone

from rbd.models import RbdHistorial, RbdServicio
from rbd.services.historial import registrar_historial_rbd


BAJA_ALIASES = {
    "zona": {"zona"},
    "rbd": {"rbd"},
    "tecnologia": {"tecnologia", "tecnología"},
    "partner": {"partner"},
    "observacion_mineduc": {"observacion mineduc", "observación mineduc"},
    "decreto": {"decreto oficio mineduc", "decreto/oficio mineduc", "oficio mineduc"},
    "fecha_notificacion": {
        "fecha notificacion de la empresa y o decreto",
        "fecha notificación de la empresa y o decreto",
        "fecha notificacion",
        "fecha notificación",
    },
    "ov": {"ov baja", "orden venta baja"},
    "observacion": {"obs", "observacion", "observación"},
}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize(value):
    text = unicodedata.normalize("NFKD", _clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[\s_./-]+", " ", text)
    return " ".join(text.split())


def _to_int(value):
    text = _clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _header_map(row):
    normalized_headers = {
        _normalize(value): index for index, value in enumerate(row) if _clean(value)
    }
    mapped = {}
    for field, aliases in BAJA_ALIASES.items():
        for alias in aliases:
            key = _normalize(alias)
            if key in normalized_headers:
                mapped[field] = normalized_headers[key]
                break
    return mapped


def _find_header_and_rows(sheet):
    iterator = sheet.iter_rows(values_only=True)
    for row in iterator:
        mapped = _header_map(row)
        if "rbd" in mapped:
            return mapped, iterator
    return {}, iter(())


def _row_value(row, mapped, field):
    index = mapped.get(field)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _worksheet(workbook):
    if "Bajas" in workbook.sheetnames:
        return workbook["Bajas"]
    return workbook.worksheets[0] if workbook.worksheets else None


def importar_bajas_excel(path, user=None):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Falta openpyxl. Reinstala dependencias del proyecto.") from exc

    source = Path(path)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = _worksheet(workbook)
    if sheet is None:
        raise ValueError("El Excel no tiene hojas disponibles.")

    mapped, rows = _find_header_and_rows(sheet)
    if "rbd" not in mapped:
        raise ValueError("No se encontro una columna RBD en el Excel de bajas.")

    baja_fecha = timezone.now()
    creados = 0
    actualizados = 0
    ya_baja = 0
    omitidos = 0

    with transaction.atomic():
        for row in rows:
            rbd = _to_int(_row_value(row, mapped, "rbd"))
            if rbd is None:
                omitidos += 1
                continue

            zona = _clean(_row_value(row, mapped, "zona"))
            tecnologia = _clean(_row_value(row, mapped, "tecnologia"))
            defaults = {
                "dado_baja": True,
                "baja_fecha": baja_fecha,
                "baja_por": user if getattr(user, "is_authenticated", False) else None,
                "baja_partner": _clean(_row_value(row, mapped, "partner")),
                "baja_observacion_mineduc": _clean(_row_value(row, mapped, "observacion_mineduc")),
                "baja_decreto": _clean(_row_value(row, mapped, "decreto")),
                "baja_notificacion_fecha": _to_date(_row_value(row, mapped, "fecha_notificacion")),
                "baja_ov": _clean(_row_value(row, mapped, "ov")),
                "baja_observacion": _clean(_row_value(row, mapped, "observacion")),
            }
            if zona:
                defaults["zona"] = zona
            if tecnologia:
                defaults["tecnologia"] = tecnologia

            servicio = RbdServicio.objects.filter(rbd=rbd).first()
            if servicio is None:
                servicio = RbdServicio.objects.create(rbd=rbd, **defaults)
                registrar_historial_rbd(
                    servicio,
                    RbdHistorial.ACCION_BAJA,
                    user,
                    "Carga masiva de Excel de bajas: servicio creado directamente como baja.",
                )
                creados += 1
                continue

            if servicio.dado_baja:
                ya_baja += 1
                registrar = False
            else:
                actualizados += 1
                registrar = True
            for field, value in defaults.items():
                setattr(servicio, field, value)
            servicio.save(update_fields=[*defaults.keys(), "actualizado_en"])
            if registrar:
                registrar_historial_rbd(
                    servicio,
                    RbdHistorial.ACCION_BAJA,
                    user,
                    "Carga masiva de Excel de bajas.",
                )

    return {
        "creados": creados,
        "actualizados": actualizados,
        "ya_baja": ya_baja,
        "omitidos": omitidos,
    }
