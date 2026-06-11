from pathlib import Path
import re
import unicodedata

from django.db import transaction

from rbd.management.commands.importar_jornada import JORNADAS
from rbd.models import RbdContacto, RbdServicio


SERVICE_COLUMNS = (
    ("rbd", "RBD"),
    ("nombre_establecimiento", "NOMBRE ESTABLECIMIENTO"),
    ("direccion", "DIRECCION"),
    ("localidad", "LOCALIDAD"),
    ("region", "REGION"),
    ("lat", "LATITUD"),
    ("long", "LONGITUD"),
    ("zona", "ZONA"),
    ("zonal", "ZONAL"),
    ("tecnologia", "TECNOLOGIA ACTUAL"),
    ("tipo", "TIPO TECNOLOGIA"),
    ("status_imaster", "STATUS IMASTER"),
    ("fiscalizacion", "FISCALIZACION"),
    ("vigencia", "VIGENCIA"),
    ("matricula", "MATRICULA"),
    ("tipo_ont", "TIPO ONT"),
    ("puerta", "PUERTA"),
    ("nodo", "NODO"),
    ("bw_nacional", "BW NACIONAL"),
    ("bw_internacional", "BW INTERNACIONAL"),
    ("bw", "BW BAJADA"),
    ("fw_usg", "FW USG"),
    ("serie_usg", "SERIE USG"),
    ("codigo_servicio_ncc", "CODIGO SERVICIO NCC"),
    ("vlan", "VLAN"),
    ("codigo_servicio_oss", "CODIGO SERVICIO OSS"),
    ("ip", "IP"),
    ("jornada_categoria", "JORNADA CATEGORIA"),
    ("jornada_horario", "JORNADA HORARIO"),
    ("jornada_descripcion", "JORNADA DESCRIPCION"),
    ("dependencia_mpls", "DEPENDENCIA MPLS"),
    ("por", "POR"),
    ("ont", "ONT"),
    ("tiene_rfs", "TIENE RFS"),
    ("rfs", "RFS"),
    ("caja", "CAJA"),
    ("fil", "FIL"),
    ("servicio_existente_oss", "SERVICIO EXISTENTE OSS"),
    ("tiene_ov", "TIENE OV"),
    ("orden_venta", "ORDEN DE VENTA"),
    ("estado_ov", "ESTADO OV"),
    ("bpi", "BPI"),
)

CONTACT_COLUMNS = (
    ("rbd", "RBD"),
    ("orden", "CONTACTO"),
    ("nombre", "NOMBRE"),
    ("telefono", "TELEFONO"),
    ("celular", "CELULAR"),
    ("email", "EMAIL"),
    ("cargo", "CARGO"),
    ("fuente", "FUENTE"),
)

NUMERIC_FIELDS = {"lat", "long", "bw_nacional", "bw_internacional", "bw"}

FIELD_ALIASES = {
    field: {field, header, header.replace(" ", "_")}
    for field, header in SERVICE_COLUMNS
}

for field, aliases in {
    "nombre_establecimiento": {
        "nombre establecimiento",
        "establecimiento ees",
        "nombre",
        "colegio",
    },
    "direccion": {"direccion ees"},
    "localidad": {"comuna"},
    "lat": {"lat"},
    "long": {"lng", "lon", "longitud"},
    "tecnologia": {"tecnologia", "tecnologia actual"},
    "tipo": {"tipo", "tipo tecnologia"},
    "bw": {"bw", "bw bajada"},
    "codigo_servicio_oss": {"codigo oss", "ef"},
    "ip": {"ip_"},
    "jornada_categoria": {"jornada", "categoria jornada"},
}.items():
    FIELD_ALIASES.setdefault(field, set()).update(aliases)

CONTACT_ALIASES = {
    field: {field, header, header.replace(" ", "_")}
    for field, header in CONTACT_COLUMNS
}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize(value):
    text = unicodedata.normalize("NFKD", _clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[_\-./]+", " ", text)
    return " ".join(text.split())


def _to_number(value):
    text = _clean(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value):
    text = _clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _header_map(row, aliases):
    normalized_headers = {
        _normalize(value): index for index, value in enumerate(row) if _clean(value)
    }
    mapped = {}
    for field, names in aliases.items():
        for alias in names:
            key = _normalize(alias)
            if key in normalized_headers:
                mapped[field] = normalized_headers[key]
                break
    return mapped


def _row_value(row, mapped, field):
    index = mapped.get(field)
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


def _find_header_and_rows(sheet, aliases, required_field="rbd"):
    iterator = sheet.iter_rows(values_only=True)
    for row in iterator:
        mapped = _header_map(row, aliases)
        if required_field in mapped:
            return mapped, iterator
    return {}, iter(())


def _worksheet(workbook, preferred_name, fallback_index=0):
    if preferred_name in workbook.sheetnames:
        return workbook[preferred_name]
    if len(workbook.worksheets) > fallback_index:
        return workbook.worksheets[fallback_index]
    return None


def _service_row(servicio):
    return [getattr(servicio, field) for field, _header in SERVICE_COLUMNS]


def _contact_row(contacto):
    return [
        contacto.servicio.rbd,
        contacto.orden,
        contacto.nombre,
        contacto.telefono,
        contacto.celular,
        contacto.email,
        contacto.cargo,
        contacto.fuente,
    ]


def _format_sheet(sheet):
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(_clean(cell.value)) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 34)


def export_carga_completa(path, include_data=True):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Falta openpyxl. Reinstala dependencias del proyecto.") from exc

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()

    servicios_sheet = workbook.active
    servicios_sheet.title = "Servicios"
    servicios_sheet.append([header for _field, header in SERVICE_COLUMNS])
    if include_data:
        for servicio in RbdServicio.objects.order_by("rbd"):
            servicios_sheet.append(_service_row(servicio))
    _format_sheet(servicios_sheet)

    contactos_sheet = workbook.create_sheet("Contactos")
    contactos_sheet.append([header for _field, header in CONTACT_COLUMNS])
    if include_data:
        contactos = RbdContacto.objects.select_related("servicio").order_by("servicio__rbd", "orden")
        for contacto in contactos:
            contactos_sheet.append(_contact_row(contacto))
    _format_sheet(contactos_sheet)

    jornada_sheet = workbook.create_sheet("Catalogo_Jornada")
    jornada_sheet.append(["CATEGORIA", "HORARIO", "DESCRIPCION"])
    for categoria, detalle in JORNADAS.items():
        jornada_sheet.append([categoria, detalle["horario"], detalle["descripcion"]])
    _format_sheet(jornada_sheet)

    instrucciones_sheet = workbook.create_sheet("Instrucciones")
    instrucciones_sheet.append(["USO", "DETALLE"])
    instrucciones_sheet.append(["Servicios", "Mantener una fila por RBD. La columna RBD es obligatoria."])
    instrucciones_sheet.append(["Contactos", "Mantener una fila por contacto. RBD y CONTACTO son obligatorios."])
    instrucciones_sheet.append(["Carga", "Subir este mismo archivo desde ADMIN > Carga de datos."])
    _format_sheet(instrucciones_sheet)

    workbook.save(target)
    return {
        "servicios": RbdServicio.objects.count() if include_data else 0,
        "contactos": RbdContacto.objects.count() if include_data else 0,
        "path": str(target),
    }


def import_carga_completa(path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Falta openpyxl. Reinstala dependencias del proyecto.") from exc

    source = Path(path)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)

    service_sheet = _worksheet(workbook, "Servicios", 0)
    if service_sheet is None:
        raise ValueError("El Excel no tiene una hoja de servicios.")

    service_map, service_rows = _find_header_and_rows(service_sheet, FIELD_ALIASES)
    if "rbd" not in service_map:
        raise ValueError("No se encontro una columna RBD en la hoja Servicios.")

    contact_sheet = _worksheet(workbook, "Contactos", 1)
    contact_map, contact_rows = ({}, iter(()))
    if contact_sheet is not None:
        contact_map, contact_rows = _find_header_and_rows(contact_sheet, CONTACT_ALIASES)

    created = 0
    updated = 0
    skipped_services = 0
    contacts_created = 0
    skipped_contacts = 0
    contact_payload = []
    contact_rbds = set()
    contact_keys = set()

    with transaction.atomic():
        for row in service_rows:
            rbd = _to_int(_row_value(row, service_map, "rbd"))
            if rbd is None:
                skipped_services += 1
                continue

            defaults = {}
            for field, _header in SERVICE_COLUMNS:
                if field == "rbd" or field not in service_map:
                    continue
                raw_value = _row_value(row, service_map, field)
                defaults[field] = _to_number(raw_value) if field in NUMERIC_FIELDS else raw_value

            categoria = (defaults.get("jornada_categoria") or "").upper()[:1]
            if categoria in JORNADAS:
                defaults["jornada_categoria"] = categoria
                if not defaults.get("jornada_horario"):
                    defaults["jornada_horario"] = JORNADAS[categoria]["horario"]
                if not defaults.get("jornada_descripcion"):
                    defaults["jornada_descripcion"] = JORNADAS[categoria]["descripcion"]

            _servicio, was_created = RbdServicio.objects.update_or_create(
                rbd=rbd,
                defaults=defaults,
            )
            created += int(was_created)
            updated += int(not was_created)

        if contact_map and "rbd" in contact_map and "orden" in contact_map:
            for row in contact_rows:
                rbd = _to_int(_row_value(row, contact_map, "rbd"))
                orden = _to_int(_row_value(row, contact_map, "orden"))
                if rbd is None or orden is None:
                    skipped_contacts += 1
                    continue
                key = (rbd, orden)
                if key in contact_keys:
                    skipped_contacts += 1
                    continue
                contact_keys.add(key)
                defaults = {
                    field: _row_value(row, contact_map, field)
                    for field, _header in CONTACT_COLUMNS
                    if field not in {"rbd", "orden"} and field in contact_map
                }
                contact_payload.append((rbd, orden, defaults))
                contact_rbds.add(rbd)

            if contact_rbds:
                RbdContacto.objects.filter(servicio__rbd__in=contact_rbds).delete()

            for rbd, orden, defaults in contact_payload:
                servicio, _was_created = RbdServicio.objects.get_or_create(rbd=rbd)
                RbdContacto.objects.create(servicio=servicio, orden=orden, **defaults)
                contacts_created += 1

    return {
        "created": created,
        "updated": updated,
        "skipped_services": skipped_services,
        "contacts": contacts_created,
        "skipped_contacts": skipped_contacts,
    }
